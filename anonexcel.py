import pandas as pd
import numpy as np
import uuid
import re
from collections import defaultdict
import random
import string
import json
import os

class DataAnonymizer:
    def __init__(self, config_file=None):
        self.mapping_dicts = defaultdict(dict)
        self.column_types = {}
        self.config_file = config_file
        self.float_precision = {}  # Для хранения точности чисел с плавающей точкой
        
        # Загружаем конфигурацию, если файл существует
        if config_file and os.path.exists(config_file):
            self.load_config()
    
    def load_config(self):
        """Загружает конфигурацию из JSON файла"""
        try:
            with open(self.config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
                self.column_types = config.get('column_types', {})
                print(f"Загружена конфигурация из {self.config_file}")
        except Exception as e:
            print(f"Ошибка загрузки конфигурации: {e}")
    
    def save_config(self, df, output_file):
        """Сохраняет конфигурацию в JSON файл"""
        config = {
            'column_types': self.column_types,
            'columns': list(df.columns)
        }
        
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
            print(f"Конфигурация сохранена в {output_file}")
        except Exception as e:
            print(f"Ошибка сохранения конфигурации: {e}")
    
    def detect_float_precision(self, value):
        """Определяет количество знаков после запятой для чисел с плавающей точкой"""
        if pd.isna(value) or not isinstance(value, (float, str)):
            return 0
        
        value_str = str(value)
        if '.' in value_str:
            decimal_part = value_str.split('.')[1]
            # Убираем trailing zeros для определения реальной точности
            return len(decimal_part.rstrip('0'))
        return 0
    
    def detect_column_type(self, series, column_name):
        """Определяет тип данных в колонке"""
        # Если тип уже задан в конфигурации, используем его
        if column_name in self.column_types:
            return self.column_types[column_name]
        
        non_null = series.dropna()
        if len(non_null) == 0:
            return 'unknown'
        
        # Проверяем на UUID (без дефисов)
        uuid_pattern = re.compile(r'^[0-9a-f]{32}$', re.I)
        uuid_samples = [str(x) for x in non_null.head(100) if pd.notna(x)]
        if all(uuid_pattern.match(x) for x in uuid_samples if x):
            return 'uuid'
        
        # Проверяем на UUID (с дефисами)
        uuid_pattern_with_dashes = re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', re.I)
        if all(uuid_pattern_with_dashes.match(str(x)) for x in non_null.head(100) if pd.notna(x)):
            return 'uuid'
        
        # Проверяем на даты
        try:
            pd.to_datetime(non_null.head(100), errors='raise')
            return 'date'
        except:
            pass
        
        # Проверяем на целые числа
        if all(isinstance(x, (int, np.integer)) or (isinstance(x, str) and x.isdigit()) for x in non_null.head(100)):
            return 'integer'
        
        # Проверяем на числа с плавающей точкой и определяем точность
        float_samples = non_null.head(100)
        if all(isinstance(x, (float, np.floating)) for x in float_samples):
            # Сохраняем максимальную точность для колонки
            max_precision = max(self.detect_float_precision(x) for x in float_samples if pd.notna(x))
            self.float_precision[column_name] = max_precision
            return 'float'
        
        # Проверяем строки, которые могут быть float
        try:
            float_values = [float(str(x).replace(',', '.')) for x in float_samples if pd.notna(x)]
            # Если все значения можно преобразовать в float
            max_precision = max(self.detect_float_precision(x) for x in float_values)
            self.float_precision[column_name] = max_precision
            return 'float'
        except:
            pass
        
        # Проверяем на алфавитно-цифровые идентификаторы
        alphanum_pattern = re.compile(r'^[A-Za-z]+[0-9]+$')
        if all(alphanum_pattern.match(str(x)) for x in non_null.head(100) if pd.notna(x)):
            return 'alphanumeric'
        
        # Проверяем на названия/текст
        word_pattern = re.compile(r'^[A-Za-z\s\.,&]+$', re.I)
        if any(word_pattern.match(str(x)) for x in non_null.head(100) if pd.notna(x)):
            return 'text'
        
        return 'unknown'
    
    def generate_random_words(self, n_words=3):
        """Генерирует случайные слова"""
        words = ['Company', 'Enterprise', 'Corporation', 'Limited', 'Group', 
                'International', 'Global', 'National', 'Regional', 'Local',
                'Solutions', 'Services', 'Technologies', 'Industries', 'Ventures']
        return ' '.join(random.sample(words, min(n_words, len(words))))
    
    def anonymize_uuid(self, value):
        """Анонимизирует UUID с сохранением маппинга (без дефисов)"""
        if pd.isna(value):
            return value
        
        if value not in self.mapping_dicts['uuid']:
            # Генерируем UUID и убираем дефисы
            new_uuid = str(uuid.uuid4()).replace('-', '')
            self.mapping_dicts['uuid'][value] = new_uuid
        return self.mapping_dicts['uuid'][value]
    
    def anonymize_integer(self, value, length=None):
        """Анонимизирует целое число"""
        if pd.isna(value):
            return value
        
        if length is None:
            length = len(str(int(value)))
        
        # Генерируем случайное число с таким же количеством цифр
        min_val = 10**(length-1)
        max_val = (10**length) - 1
        return str(random.randint(min_val, max_val))  # Возвращаем как строку
    
    def anonymize_float(self, value, column_name):
        """Анонимизирует число с плавающей точкой с сохранением формата"""
        if pd.isna(value):
            return value
        
        # Определяем количество знаков после запятой
        precision = self.float_precision.get(column_name, 0)
        
        if precision > 0:
            # Генерируем случайное число и форматируем с нужной точностью
            random_float = random.uniform(0, 1000000)
            formatted = f"{random_float:.{precision}f}"
            return formatted
        else:
            # Для целых чисел возвращаем как есть
            return str(value)
    
    def anonymize_alphanumeric(self, value):
        """Анонимизирует алфавитно-цифровые идентификаторы"""
        if pd.isna(value):
            return value
        
        value_str = str(value)
        
        # Разделяем буквенную и цифровую части
        letters = ''.join(filter(str.isalpha, value_str))
        numbers = ''.join(filter(str.isdigit, value_str))
        
        if value_str not in self.mapping_dicts['alphanumeric']:
            # Генерируем случайные цифры с тем же количеством знаков
            if numbers:
                new_numbers = str(self.anonymize_integer(numbers, len(numbers)))
            else:
                new_numbers = ''
            
            self.mapping_dicts['alphanumeric'][value_str] = letters + new_numbers
        
        return self.mapping_dicts['alphanumeric'][value_str]
    
    def anonymize_text(self, value):
        """Анонимизирует текст/названия"""
        if pd.isna(value):
            return value
        
        if value not in self.mapping_dicts['text']:
            # Определяем количество слов в оригинальном тексте
            word_count = len(str(value).split())
            self.mapping_dicts['text'][value] = self.generate_random_words(word_count)
        
        return self.mapping_dicts['text'][value]
    
    def convert_to_string(self, value):
        """Конвертирует значение в строку с сохранением формата"""
        if pd.isna(value):
            return value
        
        if isinstance(value, (int, np.integer)):
            return str(value)
        elif isinstance(value, (float, np.floating)):
            # Сохраняем оригинальный формат числа
            value_str = str(value)
            if '.' in value_str:
                # Сохраняем trailing zeros
                integer_part, decimal_part = value_str.split('.')
                decimal_part = decimal_part.rstrip('0')
                if decimal_part:
                    return f"{integer_part}.{decimal_part}"
                else:
                    return integer_part
            return value_str
        elif isinstance(value, pd.Timestamp):
            # Сохраняем дату в оригинальном формате
            return value.strftime('%Y-%m-%d')
        else:
            return str(value)
    
    def process_dataframe(self, df, generate_config=False):
        """Обрабатывает весь DataFrame"""
        result_df = df.copy()
        
        for column in df.columns:
            # Пропускаем колонки с малым количеством уникальных значений
            unique_count = df[column].nunique()
            if unique_count < 100:
                print(f"Пропускаем колонку '{column}' (словарное значение, {unique_count} уникальных значений)")
                # Все равно конвертируем в строку для единообразия
                result_df[column] = df[column].apply(self.convert_to_string)
                continue
            
            # Определяем тип колонки
            col_type = self.detect_column_type(df[column], column)
            self.column_types[column] = col_type
            print(f"Обрабатываем колонку '{column}' как {col_type}")
            
            # Применяем соответствующую анонимизацию
            if col_type == 'uuid':
                result_df[column] = df[column].apply(self.anonymize_uuid)
            elif col_type == 'integer':
                result_df[column] = df[column].apply(lambda x: self.anonymize_integer(x) if pd.notna(x) else x)
            elif col_type == 'float':
                result_df[column] = df[column].apply(lambda x: self.anonymize_float(x, column) if pd.notna(x) else x)
            elif col_type == 'alphanumeric':
                result_df[column] = df[column].apply(self.anonymize_alphanumeric)
            elif col_type == 'text':
                result_df[column] = df[column].apply(self.anonymize_text)
            elif col_type == 'date':
                # Для дат конвертируем в строку
                result_df[column] = df[column].apply(self.convert_to_string)
            else:
                # Для неизвестных типов тоже конвертируем в строку
                result_df[column] = df[column].apply(self.convert_to_string)
        
        return result_df

def save_as_text_excel(df, output_file, sheet_name='Sheet1'):
    """
    Сохраняет DataFrame в Excel с текстовым форматом всех ячеек
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import numbers
    
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Записываем данные из DataFrame
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Устанавливаем текстовый формат для всех ячеек
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            # Устанавливаем текстовый формат
            cell.number_format = '@'  # Текстовый формат в Excel
    
    # Сохраняем файл
    wb.save(output_file)

def anonymize_excel_file(input_file, output_file, config_file=None, generate_config=False, sheet_name=0):
    """
    Анонимизирует Excel файл и сохраняет все значения как текст
    
    Parameters:
    input_file (str): путь к исходному файлу
    output_file (str): путь для сохранения результата
    config_file (str): путь к файлу конфигурации
    generate_config (bool): генерировать ли файл конфигурации
    sheet_name (str/int): имя или индекс листа
    """
    
    # Читаем Excel файл
    print(f"Чтение файла: {input_file}")
    df = pd.read_excel(input_file, sheet_name=sheet_name)
    
    # Создаем анонимизатор
    anonymizer = DataAnonymizer(config_file)
    
    # Обрабатываем данные
    anonymized_df = anonymizer.process_dataframe(df, generate_config)
    
    # Сохраняем конфигурацию, если нужно
    if generate_config and config_file:
        anonymizer.save_config(df, config_file)
    
    # Сохраняем результат с текстовым форматом
    print(f"Сохранение результата в: {output_file} (все значения как текст)")
    save_as_text_excel(anonymized_df, output_file)
    
    print("Анонимизация завершена успешно!")
    print(f"Обработано колонок: {len(df.columns)}")
    print(f"Типы колонок: {anonymizer.column_types}")

# Пример использования
if __name__ == "__main__":
    # Укажите пути к вашим файлам
    input_excel = "sensitive_data.xlsx"  # Ваш исходный файл
    output_excel = "anonymized_data.xlsx"  # Файл для сохранения результата
    config_file = "column_config.json"  # Файл конфигурации
    
    # Опции
    generate_config = False  # Установите True для генерации конфигурационного файла
    
    try:
        anonymize_excel_file(
            input_excel, 
            output_excel, 
            config_file=config_file,
            generate_config=generate_config
        )
        print("Готово! Файл успешно анонимизирован.")
        print("Все значения сохранены как текст для предотвращения автоматического форматирования Excel.")
        
        if generate_config:
            print(f"Конфигурационный файл создан: {config_file}")
            print("Отредактируйте типы колонок при необходимости и запустите с generate_config=False")
            
    except Exception as e:
        print(f"Произошла ошибка: {e}")
